import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

stores = [
    (515,  "Murphy",             "NC", "lease",        26832),
    (655,  "Madisonville",       "KY", "sale",         None),
    (831,  "Albuquerque",        "NM", "sale",         None),
    (1326, "Grand Island",       "NE", "lease",        255455),
    (1739, "Washington",         "PA", "lease",        57247),
    (1824, "Ypsilanti",          "MI", "lease",        300000),
    (1983, "Ceres",              "CA", "sale",         None),
    (2305, "Bristol",            "CT", "lease",        568329),
    (2415, "Niskayuna",          "NY", "lease",        612308),
    (2666, "Columbus",           "OH", "lease",        142063),
    (2680, "South Bend",         "IN", "sale",         None),
    (2936, "Milwaukee",          "WI", "sale",         None),
    (3068, "Marietta",           "GA", "lease",        230000),
    (3116, "Marietta",           "GA", "lease",        175000),
    (3158, "Rosedale",           "NY", "lease",        875000),
    (3166, "Chicago",            "IL", "lease",        638695),
    (3186, "Alpharetta",         "GA", "lease",        310415),
    (3524, "El Cajon",           "CA", "lease",        1575625),
    (3610, "Louisville",         "KY", "sale",         None),
    (3618, "Roanoke",            "VA", "lease",        604420),
    (3656, "Cincinnati",         "OH", "sale",         None),
    (3885, "Richmond",           "VA", "lease",        661573),
    (3892, "North Chesterfield", "VA", "lease",        652938),
    (4126, "Bryan",              "TX", "lease",        297676),
    (4149, "Charlotte",          "NC", "lease",        408000),
    (4167, "Knoxville",          "TN", "lease",        245677),
    (4177, "Lincolnwood",        "IL", "lease",        912522),
    (4523, "Louisville",         "KY", "lease",        494081),
    (4661, "San Antonio",        "TX", "lease",        581302),
    (4669, "Pinellas Park",      "FL", "ground_lease", 410000),
    (4670, "San Antonio",        "TX", "lease",        573626),
    (4671, "Chattanooga",        "TN", "lease",        659572),
    (5038, "Plainfield",         "IL", "sale",         None),
    (5195, "Everett",            "WA", "sale",         None),
    (5318, "Columbia",           "SC", "lease",        623666),
    (5324, "San Antonio",        "TX", "lease",        599254),
    (5344, "Towson",             "MD", "lease",        1437921),
    (5404, "Homewood",           "IL", "lease",        600000),
    (5426, "Fremont",            "CA", "lease",        1375000),
    (5468, "Morristown",         "TN", "lease",        681495),
    (5638, "San Diego",          "CA", "lease",        389650),
    (5645, "Chicago",            "IL", "lease",        776122),
    (5646, "Chicago",            "IL", "lease",        393792),
    (5647, "Chicago",            "IL", "lease",        675180),
    (5781, "Chicago",            "IL", "sale",         None),
    (5899, "Portland",           "OR", "ground_lease", 660000),
    (5941, "Washington",         "DC", "lease",        1831200),
    (5954, "West Covina",        "CA", "lease",        1846050),
    (5980, "Granite Bay",        "CA", "lease",        624813),
    (5995, "Lake Oswego",        "OR", "lease",        272000),
    (6167, "Des Plaines",        "IL", "lease",        203485),
    (6400, "St. Petersburg",     "FL", "lease",        229780),
    (6530, "Sacramento",         "CA", "lease",        251600),
    (6574, "Virginia Beach",     "VA", "lease",        346000),
    (7310, "Katy",               "TX", "lease",        560000),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Properties"

hdr_fill = PatternFill("solid", fgColor="0053E2")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

# CRITICAL: plain machine-readable column names — no newlines, no parentheses
headers = ['city','state','property_type','listing_type','size_acres',
           'price','lat','lon','status','address','description']

ws.row_dimensions[1].height = 20
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin

widths = [20, 8, 16, 16, 12, 16, 12, 12, 16, 30, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = w

for ri, (store_num, city, state, listing, price) in enumerate(stores, 2):
    dt = "ground lease" if listing == "ground_lease" else ("sublease" if listing == "lease" else "sale")
    desc = f"Former Walmart dark store available for {dt}. Store #{store_num}."
    vals = [city, state, "buildings", listing, None, price, None, None, "available", None, desc]
    for col, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=col, value=val)
        c.alignment = Alignment(vertical="center")
        c.border = thin
        if col == 4:
            clr = {"ground_lease": "A9DDF7", "lease": "D4F0C0", "sale": "FFE8A3"}[listing]
            c.fill = PatternFill("solid", fgColor=clr)

ws.freeze_panes = "A2"
wb.save("Dark-Stores-Import-Ready.xlsx")
print(f"✅  Saved Dark-Stores-Import-Ready.xlsx | {len(stores)} rows | clean headers")
